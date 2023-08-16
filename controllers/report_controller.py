from tkinter import END

from PyScripts.views.report_view import ReportView
import openpyxl

class ReportController:
    def __init__(self, root_controller):
        self.view = ReportView(self, root_controller.view)
        self.file_name = "./BerAmer_3.xlsx"
        self.book = openpyxl.load_workbook(self.file_name)
        self.sheet_1 = self.book["Sheet1"]
        self.sheet_1.title = "Sheet1"
        self.sheet_2 = self.book["Sheet2"]
        self.sheet_2.title = "Sheet2"

    def create_spreadsheet(self):
        print("Method ran")
        #book = openpyxl.load_workbook(self.file_name)
        # Get the first sheet and give it a name
        #self.sheet_1 = self.book.active
        #self.sheet_1.title = "Sheet1"

        self.sheet_1["A2"].value = "Date"
        self.sheet_1["B2"].value = "Units Produced"
        self.sheet_1["C2"].value = "Units Sold"
        self.sheet_1["D2"].value = "Medium costs"
        self.sheet_1["E2"].value = "Container Costs"
        self.sheet_1["F2"].value = "Seed Costs"
        self.sheet_1["G2"].value = "Variable Costs"
        self.sheet_1["H2"].value = "Delivery Costs"

        self.sheet_1["B4"].value = "Expenses"
        self.sheet_1["B5"].value = "=SUM(D3, E3, F3, G3, H3)"

        self.sheet_1["C4"].value = "Revenue"
        self.sheet_1["C5"].value = "=C3*2"

        self.sheet_1["D4"].value = "Profits"
        self.sheet_1["D5"].value = "=C5-B5"

        self.sheet_1["E4"].value = "Profit/Unit Sold"

        # self.sheet_1["F5"].value = profit_per

    def updatesheet_1(self):
        #global date_entry, production_entry, sales_entry, medium_buy_entry, \
         #   container_buy_entry, seed_buy_entry, variable_buy_entry, \
          #  delivery_buy_entry, next_row

        self.sheet_1["A3"].value = self.view.date_entry.get()
        self.sheet_1["B3"].value = int(self.view.production_entry.get() or 0) + int(self.sheet_1["B3"].value or 0)
        self.sheet_1["C3"].value = int(self.view.sales_entry.get() or 0) + int(self.sheet_1["C3"].value or 0)
        self.sheet_1["D3"].value = int(self.view.medium_buy_entry.get() or 0) + int(self.sheet_1["D3"].value or 0)
        self.sheet_1["E3"].value = int(self.view.container_buy_entry.get() or 0) + int(self.sheet_1["E3"].value or 0)
        self.sheet_1["F3"].value = int(self.view.seed_buy_entry.get() or 0) + int(self.sheet_1["F3"].value or 0)
        self.sheet_1["G3"].value = int(self.view.variable_buy_entry.get() or 0) + int(self.sheet_1["G3"].value or 0)
        self.sheet_1["H3"].value = int(self.view.delivery_buy_entry.get() or 0) + int(self.sheet_1["H3"].value or 0)

    def getLastRow(self):
        rowIndex = 2

        while (True):
            value = self.sheet_2.cell(row=rowIndex, column=1).value

            if not value:
                return rowIndex

            rowIndex = rowIndex + 1

    def addNewLine(self, sheet):
        last_row = self.getLastRow()
        sheet.cell(row=last_row, column=1).value = last_row - 1
        self.mapping(sheet, last_row)

    def mapping(self, sheet, row_index):
        sheet.cell(row=row_index, column=2).value = self.view.date_entry.get()

        sheet.cell(row=row_index, column=3).value = self.view.production_entry.get() or "N/A"
        sheet.cell(row=row_index, column=4).value = self.view.sales_entry.get() or "N/A"
        sheet.cell(row=row_index, column=5).value = self.view.medium_buy_entry.get() or "N/A"
        sheet.cell(row=row_index, column=6).value = self.view.container_buy_entry.get() or "N/A"
        sheet.cell(row=row_index, column=7).value = self.view.seed_buy_entry.get() or "N/A"
        sheet.cell(row=row_index, column=8).value = self.view.variable_buy_entry.get() or "N/A"
        sheet.cell(row=row_index, column=9).value = self.view.delivery_buy_entry.get() or "N/A"

    def updateSheet_2(self):
        self.addNewLine(self.sheet_2)

    def emptyFeilds(self):
        self.view.date_entry.delete(0, END)
        self.view.production_entry.delete(0, END)
        self.view.sales_entry.delete(0, END)
        self.view.medium_buy_entry.delete(0, END)
        self.view.container_buy_entry.delete(0, END)
        self.view.seed_buy_entry.delete(0, END)
        self.view.variable_buy_entry.delete(0, END)
        self.view.delivery_buy_entry.delete(0, END)

    def update_sheets(self):
        self.create_spreadsheet()
        self.updatesheet_1()
        self.updateSheet_2()
        self.emptyFeilds()
        self.book.save(self.file_name)


