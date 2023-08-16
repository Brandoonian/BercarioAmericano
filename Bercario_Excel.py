import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
import excel
from datetime import date
from os.path import exists
from tkinter import *
from PIL import ImageTk, Image

root = Tk()
root.title("American Nursery")
root.geometry("300x300")
file_name = "./BerAmer_3.xlsx"
book = openpyxl.load_workbook(file_name)
sheet_1 = book["Sheet1"]
sheet_1.title = "Sheet1"


# Production
total_qty_produced = int(sheet_1["B3"].value or 0)
# Sales
total_units_sold = int(sheet_1["C3"].value or 0)
# Costs
total_medium_cost = int(        # R$ A generous estimate of the cost of growing medium is used for business purposes
    sheet_1["D3"].value or 0)       # Can be revised retrospectively.
total_container_cost = int(sheet_1["E3"].value or 0)  # Sale Containers (Stays with product when sold)
total_seed_cost = int(sheet_1["F3"].value or 0)  # R$
total_variable_costs = int(sheet_1["G3"].value or 0)  # Costs that are NOT reoccurring ($R)
total_delivery_costs = int(sheet_1["H3"].value or 0)

total_expenses = int(sheet_1["D3"].value or 0) + int(sheet_1["E3"].value or 0) + \
                 int(sheet_1["F3"].value or 0) + int(sheet_1["G3"].value or 0) + \
                 int(sheet_1["H3"].value or 0)
print(f"\n TOTAL COSTS: {total_expenses}")

#cost_per_plant = total_expenses / int(sheet_1["B3"].value or 0)
# print(f" COST PER PLANT: {cost_per_plant}")

PRICE = 2  # R$

total_revenue = total_units_sold * PRICE

total_profit = total_revenue - total_expenses

#profit_per = total_profit / total_qty_produced

# print(f" PROFIT PER PLANT: {profit_per_plant}")

sheet_1["A2"].value = "Date"
sheet_1["B2"].value = "Units Produced"
sheet_1["C2"].value = "Units Sold"
sheet_1["D2"].value = "Medium costs"
sheet_1["E2"].value = "Container Costs"
sheet_1["F2"].value = "Seed Costs"
sheet_1["G2"].value = "Variable Costs"
sheet_1["H2"].value = "Delivery Costs"

sheet_1["B4"].value = "Expenses"
sheet_1["B5"].value = "=SUM(D3, E3, F3, G3, H3)"

sheet_1["C4"].value = "Revenue"
sheet_1["C5"].value = "=B3*2"

sheet_1["D4"].value = "Profits"
sheet_1["D5"].value = "=C5-B5"

sheet_1["E4"].value = "Profit/Unit Sold"
print(f"{book.sheetnames}")
sheet_2 = book["Sheet2"]
sheet_2.title = "Sheet2"

sheet_2["A1"].value = "Entry #"
sheet_2["B1"].value = "Date"
sheet_2["C1"].value = "Units Produced"
sheet_2["D1"].value = "Units Sold"
sheet_2["E1"].value = "Medium costs"
sheet_2["F1"].value = "Container Costs"
sheet_2["G1"].value = "Seed Costs"
sheet_2["H1"].value = "Variable Costs"
sheet_2["I1"].value = "Delivery Costs"

def addNewLine(sheet):
    last_row = getLastRow()
    sheet.cell(row=last_row, column=1).value = last_row - 1
    mapping(sheet, last_row)

def mapping(sheet, row_index):
    sheet.cell(row=row_index, column=2).value = date_entry.get()

    sheet.cell(row=row_index, column=3).value = production_entry.get() or "N/A"
    sheet.cell(row=row_index, column=4).value = sales_entry.get() or "N/A"
    sheet.cell(row=row_index, column=5).value = medium_buy_entry.get() or "N/A"
    sheet.cell(row=row_index, column=6).value = container_buy_entry.get() or "N/A"
    sheet.cell(row=row_index, column=7).value = seed_buy_entry.get() or "N/A"
    sheet.cell(row=row_index, column=8).value = variable_buy_entry.get() or "N/A"
    sheet.cell(row=row_index, column=9).value = delivery_buy_entry.get() or "N/A"

    print(f"Date entry: {date_entry.get()}")

def getLastRow():
    rowIndex = 2

    while (True):
        value = sheet_2.cell(row=rowIndex, column=1).value

        if not value:
         return rowIndex

        rowIndex = rowIndex + 1

column_number = 1  # Specify the column number you want to check


def update_variables():
    global total_qty_produced, total_units_sold, total_medium_cost, \
        total_container_cost, total_seed_cost, total_variable_costs, \
        total_delivery_costs, total_expenses, total_revenue, total_profit

    sheet_1.title = "Sheet1"

    # Production
    total_qty_produced = int(sheet_1["B3"].value or 0)
    # Sales
    total_units_sold = int(sheet_1["C3"].value or 0)
    # Costs
    total_medium_cost = int(  # R$ A generous estimate of the cost of growing medium is used for business purposes
        sheet_1["D3"].value or 0)  # Can be revised retrospectively.
    total_container_cost = int(sheet_1["E3"].value or 0)  # Sale Containers (Stays with product when sold)
    total_seed_cost = int(sheet_1["F3"].value or 0)  # R$
    total_variable_costs = int(sheet_1["G3"].value or 0)  # Costs that are NOT reoccurring ($R)
    total_delivery_costs = int(sheet_1["H3"].value or 0)

    total_expenses = int(sheet_1["D3"].value or 0) + int(sheet_1["E3"].value or 0) + \
                     int(sheet_1["F3"].value or 0) + int(sheet_1["G3"].value or 0) + \
                     int(sheet_1["H3"].value or 0)
    print(f"\n TOTAL COSTS: {total_expenses}")

    # cost_per_plant = total_expenses / int(sheet_1["B3"].value or 0)
    # print(f" COST PER PLANT: {cost_per_plant}")

    PRICE = 2  # R$

    total_revenue = total_units_sold * PRICE

    total_profit = total_revenue - total_expenses

    # profit_per = total_profit / total_qty_produced

    # print(f" PROFIT PER PLANT: {profit_per_plant}")


def updateSheet_1():
    global date_entry, production_entry, sales_entry, medium_buy_entry, \
        container_buy_entry, seed_buy_entry, variable_buy_entry, \
        delivery_buy_entry, next_row

    sheet_1["A3"].value = date_entry.get()
    sheet_1["B3"].value = int(production_entry.get() or 0) + int(sheet_1["B3"].value or 0)
    sheet_1["C3"].value = int(sales_entry.get() or 0) + int(sheet_1["C3"].value or 0)
    sheet_1["D3"].value = int(medium_buy_entry.get() or 0) + int(sheet_1["D3"].value or 0)
    sheet_1["E3"].value = int(container_buy_entry.get() or 0) + int(sheet_1["E3"].value or 0)
    sheet_1["F3"].value = int(seed_buy_entry.get() or 0) + int(sheet_1["F3"].value or 0)
    sheet_1["G3"].value = int(variable_buy_entry.get() or 0) + int(sheet_1["G3"].value or 0)
    sheet_1["H3"].value = int(delivery_buy_entry.get() or 0) + int(sheet_1["H3"].value or 0)

def updateSheet_2():
    addNewLine(sheet_2)

def emptyFeilds():
    date_entry.delete(0, END)
    production_entry.delete(0, END)
    sales_entry.delete(0, END)
    medium_buy_entry.delete(0, END)
    container_buy_entry.delete(0, END)
    seed_buy_entry.delete(0, END)
    variable_buy_entry.delete(0, END)
    delivery_buy_entry.delete(0, END)

def updateSheets():
    create_spreadsheet()
    updateSheet_1()
    updateSheet_2()
    emptyFeilds()
    book.save(file_name)

def open_report():
    global date_entry, production_entry, sales_entry, medium_buy_entry, \
        container_buy_entry, seed_buy_entry, variable_buy_entry, \
        delivery_buy_entry

    report_win = Toplevel()
    report_win.title("Report Event")

    date_entry = Entry(report_win)
    date_entry.grid(row=0, column= 1)
    date_label = Label(report_win, text="Today's Date:")
    date_label.grid(row=0, column=0)

    production_entry = Entry(report_win)
    production_entry.grid(row=1, column=1)
    production_label = Label(report_win, text="New Production:")
    production_label.grid(row=1, column=0)

    sales_entry = Entry(report_win)
    sales_entry.grid(row=2, column=1)
    sales_label = Label(report_win, text="Sales:")
    sales_label.grid(row=2, column=0)

    medium_buy_entry = Entry(report_win)
    medium_buy_entry.grid(row=3, column=1)
    medium_label = Label(report_win, text="Medium Purchase:")
    medium_label.grid(row=3, column=0)

    container_buy_entry = Entry(report_win)
    container_buy_entry.grid(row=4, column=1)
    container_label = Label(report_win, text="Container Purchase:")
    container_label.grid(row=4, column=0)

    seed_buy_entry = Entry(report_win)
    seed_buy_entry.grid(row=5, column=1)
    seed_label = Label(report_win, text="Seed Purchase:")
    seed_label.grid(row=5, column=0)

    variable_buy_entry = Entry(report_win)
    variable_buy_entry.grid(row=6, column=1)
    variable_label = Label(report_win, text="Variable Purchase:")
    variable_label.grid(row=6, column=0)

    delivery_buy_entry = Entry(report_win)
    delivery_buy_entry.grid(row=7, column=1)
    delivery_label = Label(report_win, text="Delivery Fee:")
    delivery_label.grid(row=7, column=0)

    # Create 'Submit' button
    sub_butt = Button(report_win, text="Submit", command=updateSheets)
    sub_butt.grid(row=7, column=1)

    back_butt = Button(report_win, text="Back", command=report_win.destroy)
    back_butt.grid(row=7, column=2)

def open_view():
    update_variables()

    view_win = Toplevel()
    view_win.title("Expenses, Revenue, Profit")

    expenses_label = Label(view_win, text="EXPENSES:    ")
    expenses_label.grid(row=0, column=0)
    expenses_num = Label(view_win, text=total_expenses)
    expenses_num.grid(row=0, column=1)

    revenue_label = Label(view_win, text="REVENUE:    ")
    revenue_label.grid(row=1, column=0)
    revenue_num = Label(view_win, text=total_revenue)
    revenue_num.grid(row=1, column=1)

    profit_label = Label(view_win, text="PROFIT:    ")
    profit_label.grid(row=2, column=0)
    profit_num = Label(view_win, text=total_profit)
    profit_num.grid(row=2, column=1)

    profit_per_label = Label(view_win, text="PROFIT/UNIT SOLD:    ")
    profit_per_label.grid(row=3, column=0)
#    profit_per_num = Label(view_win, text=profit_per)
 #   profit_per_num.grid(row=3, column=1)

    back_butt = Button(view_win, text="Back", command=view_win.destroy)
    back_butt.grid(row=4, column=1)


class Table:

    def __init__(self, root, lst):

        # code for creating table
        for i in range(total_rows):
            for j in range(total_columns):
                self.e = Entry(root, width=10, fg='blue',
                               font=('Arial', 16, 'bold'))

                self.e.grid(row=i, column=j)
                self.e.insert(END, lst[i][j] or "")
                print(lst[i][j])

last_ten = []
table = []
get_last_rows = range((getLastRow() - 10), (getLastRow()))
for row in get_last_rows:
    index = row
    table.append([sheet_2.cell(row=index, column=2).value,
             sheet_2.cell(row=index, column=3).value, sheet_2.cell(row=index, column=4).value,
             sheet_2.cell(row=index, column=5).value, sheet_2.cell(row=index, column=6).value,
             sheet_2.cell(row=index, column=7).value, sheet_2.cell(row=index, column=8).value,
             sheet_2.cell(row=index, column=9).value])
    last_ten.append(row)


# take the data
lst = table

# find total number of rows and
# columns in list
total_rows = len(lst)
total_columns = len(lst[0])

table_win = Toplevel(root)
t = Table(table_win, lst)
table_win.title("Table")
table_win.geometry("1000x1000")


def create_spreadsheet():
    print("Method ran")
    book = openpyxl.load_workbook(file_name)
    # Get the first sheet and give it a name
    sheet_1 = book.active
    sheet_1.title = "Sheet1"

    sheet_1["A2"].value = "Date"
    sheet_1["B2"].value = "Units Produced"
    sheet_1["C2"].value = "Units Sold"
    sheet_1["D2"].value = "Medium costs"
    sheet_1["E2"].value = "Container Costs"
    sheet_1["F2"].value = "Seed Costs"
    sheet_1["G2"].value = "Variable Costs"
    sheet_1["H2"].value = "Delivery Costs"

    sheet_1["B4"].value = "Expenses"
    sheet_1["B5"].value = "=SUM(D3, E3, F3, G3, H3)"

    sheet_1["C4"].value = "Revenue"
    sheet_1["C5"].value = "=B3*2"

    sheet_1["D4"].value = "Profits"
    sheet_1["D5"].value = "=C5-B5"

    sheet_1["E4"].value = "Profit/Unit Sold"
    
    # sheet_1["F5"].value = profit_per

home_label = Label(root, text="Home")
home_label.grid(row=0, column=2)

report_butt = Button(root, text="Report Sale, Production, or Expense", command=open_report)
report_butt.grid(row=1, column=2)

view_butt = Button(root, text="View Expenses, Revenue, and Profit", command=open_view)
view_butt.grid(row=2, column=2)




root.mainloop()



