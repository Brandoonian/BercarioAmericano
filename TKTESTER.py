# Python program to create a table

from tkinter import *
import openpyxl
file_name = "./BerAmer_3.xlsx"
book = openpyxl.load_workbook(file_name)
sheet_2 = book["Sheet2"]

class Table:

    def __init__(self, root):

        # code for creating table
        for i in range(total_rows):
            for j in range(total_columns):
                self.e = Entry(root, width=20, fg='blue',
                               font=('Arial', 16, 'bold'))

                self.e.grid(row=i, column=j)
                self.e.insert(END, lst[i][j])


# take the data
lst = [sheet_2.cell(row=last_ten[0], column=2).value,
                 sheet_2.cell(row=last_ten[0], column=3).value, sheet_2.cell(row=last_ten[0], column=4).value,
                 sheet_2.cell(row=last_ten[0], column=5).value, sheet_2.cell(row=last_ten[0], column=6).value,
                 sheet_2.cell(row=last_ten[0], column=7).value, sheet_2.cell(row=last_ten[0], column=8).value,
                 sheet_2.cell(row=last_ten[0], column=9).value],\
            [sheet_2.cell(row=last_ten[1], column=2).value,
             sheet_2.cell(row=last_ten[1], column=3).value, sheet_2.cell(row=last_ten[1], column=4).value,
             sheet_2.cell(row=last_ten[1], column=5).value, sheet_2.cell(row=last_ten[1], column=6).value,
             sheet_2.cell(row=last_ten[1], column=7).value, sheet_2.cell(row=last_ten[1], column=8).value,
             sheet_2.cell(row=last_ten[1], column=9).value],\
            [sheet_2.cell(row=last_ten[2], column=2).value,
             sheet_2.cell(row=last_ten[2], column=3).value, sheet_2.cell(row=last_ten[2], column=4).value,
             sheet_2.cell(row=last_ten[2], column=5).value, sheet_2.cell(row=last_ten[2], column=6).value,
             sheet_2.cell(row=last_ten[2], column=7).value, sheet_2.cell(row=last_ten[2], column=8).value,
             sheet_2.cell(row=last_ten[2], column=9).value], \
            [sheet_2.cell(row=last_ten[3], column=2).value,
             sheet_2.cell(row=last_ten[3], column=3).value, sheet_2.cell(row=last_ten[3], column=4).value,
             sheet_2.cell(row=last_ten[3], column=5).value, sheet_2.cell(row=last_ten[3], column=6).value,
             sheet_2.cell(row=last_ten[3], column=7).value, sheet_2.cell(row=last_ten[3], column=8).value,
             sheet_2.cell(row=last_ten[3], column=9).value], \
            [sheet_2.cell(row=last_ten[4], column=2).value,
             sheet_2.cell(row=last_ten[4], column=3).value, sheet_2.cell(row=last_ten[4], column=4).value,
             sheet_2.cell(row=last_ten[4], column=5).value, sheet_2.cell(row=last_ten[4], column=6).value,
             sheet_2.cell(row=last_ten[4], column=7).value, sheet_2.cell(row=last_ten[4], column=8).value,
             sheet_2.cell(row=last_ten[4], column=9).value],\
            [sheet_2.cell(row=last_ten[5], column=2).value,
             sheet_2.cell(row=last_ten[5], column=3).value, sheet_2.cell(row=last_ten[5], column=4).value,
             sheet_2.cell(row=last_ten[5], column=5).value, sheet_2.cell(row=last_ten[5], column=6).value,
             sheet_2.cell(row=last_ten[5], column=7).value, sheet_2.cell(row=last_ten[5], column=8).value,
             sheet_2.cell(row=last_ten[5], column=9).value],\
            [sheet_2.cell(row=last_ten[6], column=2).value,
             sheet_2.cell(row=last_ten[6], column=3).value, sheet_2.cell(row=last_ten[6], column=4).value,
             sheet_2.cell(row=last_ten[6], column=5).value, sheet_2.cell(row=last_ten[6], column=6).value,
             sheet_2.cell(row=last_ten[6], column=7).value, sheet_2.cell(row=last_ten[6], column=8).value,
             sheet_2.cell(row=last_ten[6], column=9).value],\
            [sheet_2.cell(row=last_ten[7], column=2).value,
             sheet_2.cell(row=last_ten[7], column=3).value, sheet_2.cell(row=last_ten[7], column=4).value,
             sheet_2.cell(row=last_ten[7], column=5).value, sheet_2.cell(row=last_ten[7], column=6).value,
             sheet_2.cell(row=last_ten[7], column=7).value, sheet_2.cell(row=last_ten[7], column=8).value,
             sheet_2.cell(row=last_ten[7], column=9).value],\
            [sheet_2.cell(row=last_ten[8], column=2).value,
             sheet_2.cell(row=last_ten[8], column=3).value, sheet_2.cell(row=last_ten[8], column=4).value,
             sheet_2.cell(row=last_ten[8], column=5).value, sheet_2.cell(row=last_ten[8], column=6).value,
             sheet_2.cell(row=last_ten[8], column=7).value, sheet_2.cell(row=last_ten[8], column=8).value,
             sheet_2.cell(row=last_ten[8], column=9).value],\
            [sheet_2.cell(row=last_ten[9], column=2).value,
             sheet_2.cell(row=last_ten[9], column=3).value, sheet_2.cell(row=last_ten[9], column=4).value,
             sheet_2.cell(row=last_ten[9], column=5).value, sheet_2.cell(row=last_ten[9], column=6).value,
             sheet_2.cell(row=last_ten[9], column=7).value, sheet_2.cell(row=last_ten[9], column=8).value,
             sheet_2.cell(row=last_ten[9], column=9).value]

# find total number of rows and
# columns in list
total_rows = len(lst)
total_columns = len(lst[0])

# create root window
root = Tk()
t = Table(root)
root.mainloop()